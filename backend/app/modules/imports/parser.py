"""Parsers de CSV, XLSX y PDF para importación de transacciones.

Devuelven una secuencia de filas como diccionarios `{columna: valor}`.
La detección de formato es por extensión y mime-type. Los valores se
normalizan a `str` para que la lógica de validación posterior los
parsee homogéneamente.

PDF: usa `pdfplumber.extract_tables()` con la heurística por defecto.
PDFs sin texto extraíble (escaneados) terminan en `ParseError`.
"""

from __future__ import annotations

import csv
import io
from typing import IO, Any

import pdfplumber
from openpyxl import load_workbook

# Mime-types aceptados. Algunos clientes envían `application/octet-stream`,
# en cuyo caso se decide por extensión.
CSV_MIME_TYPES = {
    "text/csv",
    "application/csv",
    "application/vnd.ms-excel",  # algunos navegadores lo etiquetan así
}
XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet",
}
PDF_MIME_TYPES = {"application/pdf"}


class ParseError(Exception):
    """Error al parsear un fichero subido."""


def detect_format(filename: str, content_type: str | None) -> str:
    """Devuelve `csv`, `xlsx` o `pdf`, o lanza `ParseError` si no se reconoce."""
    name = filename.lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".pdf"):
        return "pdf"
    if content_type in CSV_MIME_TYPES:
        return "csv"
    if content_type in XLSX_MIME_TYPES:
        return "xlsx"
    if content_type in PDF_MIME_TYPES:
        return "pdf"
    raise ParseError(f"Formato no soportado: {filename!r} (content-type={content_type!r})")


def parse_csv(payload: bytes) -> list[dict[str, str]]:
    """Parsea un CSV. Detecta delimitador (`,`, `;`, `\\t`)."""
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = payload.decode("latin-1")
        except UnicodeDecodeError as e:
            raise ParseError("No se pudo decodificar el fichero (UTF-8 ni Latin-1)") from e

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ParseError("El CSV no tiene cabecera")

    rows: list[dict[str, str]] = []
    for raw in reader:
        row = {(k or "").strip(): (v or "").strip() for k, v in raw.items() if k is not None}
        rows.append(row)
    return rows


def parse_xlsx(payload: bytes) -> list[dict[str, str]]:
    """Parsea la primera hoja de un XLSX."""
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"XLSX inválido: {e}") from e

    sheet = workbook.active
    if sheet is None:
        raise ParseError("El XLSX no contiene hojas")

    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as e:
        raise ParseError("El XLSX está vacío") from e

    headers = [_stringify(cell).strip() for cell in header_row]
    if not any(headers):
        raise ParseError("El XLSX no tiene cabecera")

    rows: list[dict[str, str]] = []
    for raw_row in iterator:
        if all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue
        row = {
            headers[i]: _stringify(value).strip()
            for i, value in enumerate(raw_row)
            if i < len(headers) and headers[i]
        }
        rows.append(row)
    return rows


def parse_pdf(payload: bytes) -> list[dict[str, str]]:
    """Parsea tablas de un PDF con texto extraíble.

    Usa la primera tabla detectada como referencia: su primera fila es la
    cabecera global. Tablas en páginas sucesivas (típico en extractos
    bancarios) se concatenan; si la primera fila se repite la salta para
    evitar duplicados de cabecera.

    Limitaciones:
    - PDFs sin capa de texto (escaneados) → `ParseError`. La opción de
      fallback con visión local se evaluará tras PHASE-5.1.
    - Si las tablas tienen distinto número de columnas, se mapean por
      posición hasta el menor de los dos tamaños.
    """
    try:
        pdf = pdfplumber.open(io.BytesIO(payload))
    except Exception as e:
        raise ParseError(f"PDF inválido: {e}") from e

    tables: list[list[list[str | None]]] = []
    with pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if table:
                    tables.append(table)

    if not tables:
        raise ParseError("No se detectaron tablas en el PDF (¿escaneado?)")

    first_header_raw = tables[0][0]
    if not first_header_raw:
        raise ParseError("La primera tabla del PDF no tiene cabecera")

    header_cleaned = [_pdf_clean(cell) for cell in first_header_raw]
    headers = [
        cell or f"col_{idx}" for idx, cell in enumerate(header_cleaned)
    ]
    if not any(header_cleaned):
        raise ParseError("La cabecera del PDF está vacía")

    rows: list[dict[str, str]] = []
    for table in tables:
        for raw_row in table:
            cleaned = [_pdf_clean(c) for c in raw_row]
            if all(not c for c in cleaned):
                continue
            if cleaned[: len(header_cleaned)] == header_cleaned:
                # cabecera repetida en cabecera de página
                continue
            mapped: dict[str, str] = {}
            for i, value in enumerate(cleaned):
                if i >= len(headers):
                    break
                key = headers[i]
                if key:
                    mapped[key] = value
            rows.append(mapped)

    if not rows:
        raise ParseError("La tabla del PDF está vacía")

    return rows


def _pdf_clean(value: object | None) -> str:
    """Normaliza una celda del PDF: elimina saltos internos y espacios extra."""
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    return " ".join(text.split())


def parse_file(payload: bytes, filename: str, content_type: str | None) -> list[dict[str, str]]:
    """Parsea según formato detectado."""
    fmt = detect_format(filename, content_type)
    if fmt == "csv":
        return parse_csv(payload)
    if fmt == "xlsx":
        return parse_xlsx(payload)
    return parse_pdf(payload)


def parse_stream(stream: IO[bytes], filename: str, content_type: str | None) -> list[dict[str, str]]:
    """Variante para streams ya abiertos (p.ej. `UploadFile`)."""
    return parse_file(stream.read(), filename, content_type)


def _stringify(value: Any) -> str:
    """Convierte un valor de openpyxl a string preservando ISO en datetimes."""
    if value is None:
        return ""
    # openpyxl ya devuelve datetime/date/Decimal/int/float — basta str()
    return str(value)
