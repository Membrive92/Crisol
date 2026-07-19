"""Parsers de CSV, XLSX y PDF para importación de transacciones.

Devuelven una secuencia de filas como diccionarios `{columna: valor}`.
La detección de formato es por extensión y mime-type. Los valores se
normalizan a `str` para que la lógica de validación posterior los
parsee homogéneamente.

PDF: usa `pdfplumber.extract_tables()` con la heurística por defecto.
Los PDFs sin tablas extraíbles (escaneados) lanzan `NoTablesInPdfError`,
que el service de imports captura para fallback con visión local.

`parse_pdf_smart` añade una capa heurística sobre `parse_pdf` para
extractos bancarios reales (varias tablas: resumen + desglose +
movimientos). Detecta la tabla de transacciones por cabecera y
devuelve filas con keys fijas (`amount`, `occurred_at`, `description`,
`category_name`). Si la heurística no es confiable lanza
`SmartParseAmbiguous` y el caller cae al parser legacy.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date
from typing import Any

import pdfplumber
from openpyxl import load_workbook

# Mime-types aceptados. Algunos clientes envían `application/octet-stream`,
# en cuyo caso se decide por extensión.
CSV_MIME_TYPES = {
    "text/csv",
    "application/csv",
    "application/vnd.ms-excel",  # algunos navegadores lo etiquetan así
}
XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet",
}
PDF_MIME_TYPES = {"application/pdf"}


class ParseError(Exception):
    """Error al parsear un fichero subido."""


class NoTablesInPdfError(ParseError):
    """El PDF no tiene tablas extraíbles (típicamente escaneado).

    El service de imports lo captura para intentar el fallback con visión
    local (renderizar páginas como imagen y pasarlas al modelo).
    """


class SmartParseAmbiguousError(ParseError):
    """`parse_pdf_smart` no logró identificar con confianza la tabla
    de transacciones. El caller debe caer al parser legacy con mapping
    manual o al fallback de visión.
    """


def detect_format(filename: str, content_type: str | None) -> str:
    """Devuelve `csv`, `xlsx` o `pdf`, o lanza `ParseError` si no se reconoce."""
    name = filename.lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".pdf"):
        return "pdf"
    if content_type in CSV_MIME_TYPES:
        return "csv"
    if content_type in XLSX_MIME_TYPES:
        return "xlsx"
    if content_type in PDF_MIME_TYPES:
        return "pdf"
    raise ParseError(f"Formato no soportado: {filename!r} (content-type={content_type!r})")


def parse_csv(payload: bytes) -> list[dict[str, str]]:
    """Parsea un CSV. Detecta delimitador (`,`, `;`, `\\t`)."""
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = payload.decode("latin-1")
        except UnicodeDecodeError as e:
            raise ParseError("No se pudo decodificar el fichero (UTF-8 ni Latin-1)") from e

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ParseError("El CSV no tiene cabecera")

    rows: list[dict[str, str]] = []
    for raw in reader:
        row = {(k or "").strip(): (v or "").strip() for k, v in raw.items() if k is not None}
        rows.append(row)
    return rows


def parse_xlsx(payload: bytes) -> list[dict[str, str]]:
    """Parsea la primera hoja de un XLSX.

    Tolera bloques de cabecera bancarios típicos: muchos extractos
    (BBVA, Santander, ING, CaixaBank…) llevan 5-10 filas iniciales
    con logo, periodo, saldo inicial, etc., y la cabecera real de la
    tabla (Fecha | Concepto | Importe | Saldo) aparece más abajo.

    Heurística: escanea hasta `_XLSX_HEADER_SCAN_LIMIT` filas y elige
    como cabecera la primera con ≥2 celdas no vacías (un título
    suelto en una sola columna no se confunde con cabecera). El
    resto se procesa como datos a partir de ahí.
    """
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"XLSX inválido: {e}") from e

    sheet = workbook.active
    if sheet is None:
        raise ParseError("El XLSX no contiene hojas")

    iterator = sheet.iter_rows(values_only=True)
    headers: list[str] = []
    for rows_scanned, raw_row in enumerate(iterator, start=1):
        candidate = [_stringify(cell).strip() for cell in raw_row]
        non_empty = sum(1 for cell in candidate if cell)
        if non_empty >= 2:
            headers = candidate
            break
        if rows_scanned >= _XLSX_HEADER_SCAN_LIMIT:
            break

    if not headers:
        raise ParseError(
            "El XLSX no tiene cabecera detectable en las primeras "
            f"{_XLSX_HEADER_SCAN_LIMIT} filas"
        )

    rows: list[dict[str, str]] = []
    for raw_row in iterator:
        if all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue
        row = {
            headers[i]: _stringify(value).strip()
            for i, value in enumerate(raw_row)
            if i < len(headers) and headers[i]
        }
        rows.append(row)
    return rows


_XLSX_HEADER_SCAN_LIMIT = 30
# Cota anti decompression-bomb: un extracto bancario real rara vez supera unas
# decenas de páginas; > 50 lo tratamos como abusivo (AUDIT-2026-05).
_MAX_PDF_PAGES = 50


def parse_xlsx_smart(payload: bytes) -> list[dict[str, str]]:
    """Espejo de `parse_pdf_smart` para XLSX: detecta los roles de las
    columnas en lugar de pedir al usuario que las mapee.

    Justificación: los XLSX de bancos españoles (BBVA, Santander, ING,
    CaixaBank…) traen una sola columna "Concepto" con la descripción
    del movimiento. El wizard de mapping del frontend la asigna a
    `description` por defecto, dejando `category_name` vacío — y sin
    `category_name` el backend no construye `bank_concept_groups`,
    así que el autocompletado de categorías nunca se activa para XLSX.

    Esta función reusa la heurística del PDF smart (`_classify_columns`,
    `_AMOUNT_HEADER_HINTS`, etc.) para producir filas con keys fijas
    `amount`, `occurred_at`, `description`, `category_name`. Cuando
    sólo existe una columna tipo "Concepto", se copia su valor a las
    dos (description + category_name) — mismo truco que el PDF smart.

    Lanza `SmartParseAmbiguous` si no logra identificar columnas de
    importe Y fecha; el caller cae al `parse_xlsx` legacy con el
    mapping del usuario.
    """
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"XLSX inválido: {e}") from e

    sheet = workbook.active
    if sheet is None:
        raise ParseError("El XLSX no contiene hojas")

    iterator = sheet.iter_rows(values_only=True)
    headers: list[str] = []
    for rows_scanned, raw_row in enumerate(iterator, start=1):
        candidate = [_stringify(cell).strip() for cell in raw_row]
        non_empty = sum(1 for cell in candidate if cell)
        if non_empty >= 2:
            headers = candidate
            break
        if rows_scanned >= _XLSX_HEADER_SCAN_LIMIT:
            break

    if not headers:
        raise SmartParseAmbiguousError(
            "Sin cabecera detectable en las primeras " f"{_XLSX_HEADER_SCAN_LIMIT} filas"
        )

    roles = _classify_columns(headers)
    if "amount" not in roles or "occurred_at" not in roles:
        # Sin columnas críticas no podemos seguir — el legacy parser
        # con mapping manual del usuario lo arreglará.
        raise SmartParseAmbiguousError("No se reconocieron columnas de importe o fecha")

    rows: list[dict[str, str]] = []
    for raw_row in iterator:
        if all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue
        cleaned = [_stringify(cell).strip() for cell in raw_row]

        def get(role: str, cleaned: list[str] = cleaned) -> str:
            idx = roles.get(role)
            if idx is None or idx >= len(cleaned):
                return ""
            return cleaned[idx]

        description = get("description")
        category_name = get("category_name")
        # Mismo post-proceso que parse_pdf_smart: si no hay description,
        # copiamos category_name; si los dos son idénticos, mantenemos
        # description sola para no duplicar al usuario.
        if not description:
            description = category_name
        elif description == category_name:
            category_name = ""

        rows.append(
            {
                "amount": get("amount"),
                "occurred_at": get("occurred_at"),
                "description": description,
                "category_name": category_name,
                # PHASE-39 — saldo del extracto (vacío si no hay columna).
                "statement_balance": get("statement_balance"),
            }
        )

    if not rows:
        raise SmartParseAmbiguousError("Cabecera detectada pero sin filas de datos")

    return rows


def parse_pdf(payload: bytes) -> list[dict[str, str]]:
    """Parsea tablas de un PDF con texto extraíble.

    Usa la primera tabla detectada como referencia: su primera fila es la
    cabecera global. Tablas en páginas sucesivas (típico en extractos
    bancarios) se concatenan; si la primera fila se repite la salta para
    evitar duplicados de cabecera.

    Limitaciones:
    - PDFs sin capa de texto (escaneados) → `ParseError`. La opción de
      fallback con visión local se evaluará tras PHASE-5.1.
    - Si las tablas tienen distinto número de columnas, se mapean por
      posición hasta el menor de los dos tamaños.
    """
    try:
        pdf = pdfplumber.open(io.BytesIO(payload))
    except Exception as e:
        # `str(e)` viene vacío para varias excepciones de PDFs cifrados
        # / corruptos. Caemos al tipo para que el usuario tenga al
        # menos una pista accionable en el log.
        detail = str(e) or f"{type(e).__name__} sin mensaje (probablemente PDF cifrado o corrupto)"
        raise ParseError(f"PDF inválido: {detail}") from e

    tables: list[list[list[str | None]]] = []
    with pdf:
        if len(pdf.pages) > _MAX_PDF_PAGES:
            raise ParseError(f"PDF demasiado largo (> {_MAX_PDF_PAGES} páginas).")
        for page in pdf.pages:
            for table in page.extract_tables():
                if table:
                    tables.append(table)

    if not tables:
        raise NoTablesInPdfError("No se detectaron tablas en el PDF (¿escaneado?)")

    first_header_raw = tables[0][0]
    if not first_header_raw:
        raise ParseError("La primera tabla del PDF no tiene cabecera")

    header_cleaned = [_pdf_clean(cell) for cell in first_header_raw]
    headers = [cell or f"col_{idx}" for idx, cell in enumerate(header_cleaned)]
    if not any(header_cleaned):
        raise ParseError("La cabecera del PDF está vacía")

    rows: list[dict[str, str]] = []
    for table in tables:
        for raw_row in table:
            cleaned = [_pdf_clean(c) for c in raw_row]
            if all(not c for c in cleaned):
                continue
            if cleaned[: len(header_cleaned)] == header_cleaned:
                # cabecera repetida en cabecera de página
                continue
            mapped: dict[str, str] = {}
            for i, value in enumerate(cleaned):
                if i >= len(headers):
                    break
                key = headers[i]
                if key:
                    mapped[key] = value
            rows.append(mapped)

    if not rows:
        raise ParseError("La tabla del PDF está vacía")

    return rows


def _pdf_clean(value: object | None) -> str:
    """Normaliza una celda del PDF: elimina saltos internos y espacios extra."""
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    return " ".join(text.split())


def parse_file(payload: bytes, filename: str, content_type: str | None) -> list[dict[str, str]]:
    """Parsea según formato detectado."""
    fmt = detect_format(filename, content_type)
    if fmt == "csv":
        return parse_csv(payload)
    if fmt == "xlsx":
        return parse_xlsx(payload)
    return parse_pdf(payload)


def count_pdf_pages(payload: bytes) -> int:
    """Número total de páginas de un PDF.

    AUDIT-2026-07 (M-04): lo usa el fallback de visión para DETECTAR que un
    extracto excede el tope de páginas renderizables y avisar en vez de
    truncar en silencio (importar sólo las primeras N y reportar éxito).
    """
    try:
        import pypdfium2 as pdfium  # type: ignore[import-untyped]
    except ImportError as e:
        raise ParseError("pypdfium2 no disponible") from e
    try:
        pdf = pdfium.PdfDocument(payload)
    except Exception as e:
        raise ParseError(f"PDF inválido para render: {e}") from e
    try:
        return len(pdf)
    finally:
        pdf.close()


def render_pdf_pages_to_png(payload: bytes, *, max_pages: int, dpi: int = 150) -> list[bytes]:
    """Renderiza las primeras `max_pages` páginas del PDF a PNG.

    Se usa en el fallback de visión cuando `pdfplumber` no detecta tablas.
    Limitamos el número de páginas para evitar tiempos de inferencia
    desbordados — extractos típicos de banca tienen 1-3 páginas relevantes.
    """
    if max_pages <= 0:
        return []
    try:
        import pypdfium2 as pdfium
    except ImportError as e:
        raise ParseError("pypdfium2 no disponible") from e

    try:
        pdf = pdfium.PdfDocument(payload)
    except Exception as e:
        raise ParseError(f"PDF inválido para render: {e}") from e

    pages: list[bytes] = []
    try:
        scale = dpi / 72  # PDF default es 72 DPI.
        for index in range(min(len(pdf), max_pages)):
            page = pdf[index]
            try:
                bitmap = page.render(scale=scale)
                try:
                    image = bitmap.to_pil()
                    buffer = io.BytesIO()
                    image.save(buffer, format="PNG")
                    pages.append(buffer.getvalue())
                finally:
                    bitmap.close()
            finally:
                page.close()
    finally:
        pdf.close()
    return pages


def _stringify(value: Any) -> str:
    """Convierte un valor de openpyxl a string preservando ISO en datetimes."""
    if value is None:
        return ""
    # openpyxl ya devuelve datetime/date/Decimal/int/float — basta str()
    return str(value)


# ─────────────────────────────────────────────────────────────────────────────
# Smart PDF parser: detecta la tabla de transacciones y devuelve filas con
# keys fijas. Específico para extractos bancarios con múltiples tablas
# (resumen, desglose, movimientos). Si la heurística no es confiable, lanza
# SmartParseAmbiguous para que el caller decida el fallback.
# ─────────────────────────────────────────────────────────────────────────────

# Sinónimos de cada rol de columna (case + acento-insensibles).
_DATE_HEADER_HINTS = (
    "fecha",
    "fec.",
    "f. operac",
    "f.operac",
    "f operac",
    "f. contab",
    "f.contab",
    "f contab",
    "operación",
    "operacion",
    "fecha valor",
    "f. valor",
    "f.valor",
    "date",
)
# IMPORTANT: NO incluir "saldo" — los extractos típicos tienen
# columnas Importe + Saldo y la heurística de "último match" se quedaría
# con el saldo (acumulado), no con el importe de la transacción. El
# usuario verá saldos como si fueran transacciones — bug grave.
# AUDIT finding #7: NO incluir "valor" suelto. El comentario decía que no
# debía estar, pero el hint seguía en la lista — "valor" hacía match con
# cualquier cabecera que lo contuviera (p. ej. "Valoración"). Para
# soportar "Fecha valor" usamos hints específicos ("fecha valor",
# "f. valor") en lugar del genérico. Nota: "occurred_at" se asigna al
# PRIMER match de fecha (`_classify_columns`), así que "Fecha" (operación)
# gana sobre "Fecha valor" cuando ambas existen — el orden de la cabecera
# decide, no el orden de estos hints.
_AMOUNT_HEADER_HINTS = (
    "importe",
    "amount",
    "monto",
    "cantidad",
    "cuantia",
    "cuantía",
)
_CATEGORY_HEADER_HINTS = (
    "concepto",
    "categoría",
    "categoria",
    "tipo",
    "movimiento",
    "operación",
    "operacion",
    "category",
)
_DESCRIPTION_HEADER_HINTS = (
    "detalle",
    "descripción",
    "descripcion",
    "observaciones",
    "comercio",
    "description",
    "referencia",
)
# PHASE-39 — columna Saldo del extracto (saldo de la cuenta TRAS cada
# movimiento). Es un rol PROPIO, nunca `amount` (el IMPORTANT de arriba
# sigue vigente: "saldo" jamás debe matchear importe). Se captura para
# anclar el saldo real de la cuenta al que declara el banco.
_BALANCE_HEADER_HINTS = (
    "saldo",
    "balance",
)

# Score mínimo para aceptar una tabla como "tabla de transacciones".
_SMART_MIN_SCORE = 8


def _norm(value: str) -> str:
    """Normaliza para comparar cabeceras (lowercase + sin acentos + trim)."""
    return (
        value.lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
        .strip()
    )


def _header_matches(cell: str, hints: tuple[str, ...]) -> bool:
    n = _norm(cell)
    return any(hint in n for hint in hints)


def _classify_columns(headers: list[str]) -> dict[str, int]:
    """Devuelve `{role: column_idx}` para los roles detectados.

    Si hay dos columnas de fecha (F. Operac. + F. Contab.), gana la primera
    (típicamente la de operación, que es la que el usuario quiere). Para
    importe gana la última (suelen poner saldo antes de importe en algunos
    extractos, pero el más a la derecha tiende a ser el importe firmado).
    """
    roles: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        if "occurred_at" not in roles and _header_matches(header, _DATE_HEADER_HINTS):
            roles["occurred_at"] = idx
            continue
        if _header_matches(header, _AMOUNT_HEADER_HINTS):
            # gana el último match
            roles["amount"] = idx
            continue
        # PHASE-39 — Saldo del extracto: primer match, rol independiente.
        # No compite con `amount` (los hints no se solapan por diseño).
        if "statement_balance" not in roles and _header_matches(header, _BALANCE_HEADER_HINTS):
            roles["statement_balance"] = idx
            continue
        if "description" not in roles and _header_matches(header, _DESCRIPTION_HEADER_HINTS):
            roles["description"] = idx
            continue
        if "category_name" not in roles and _header_matches(header, _CATEGORY_HEADER_HINTS):
            roles["category_name"] = idx
    return roles


def _score_table(table: list[list[str | None]]) -> tuple[int, dict[str, int]]:
    """Devuelve `(score, role_columns)` para una tabla candidata.

    Score basis:
      +5 si la cabecera tiene una columna de fecha
      +5 si la cabecera tiene una columna de importe
      +3 si tiene columna de descripción
      +2 si tiene columna de categoría
      +1 por cada 5 filas de datos (capado a +5)
    """
    if not table or not table[0]:
        return 0, {}
    headers = [_pdf_clean(c) for c in table[0]]
    roles = _classify_columns(headers)

    score = 0
    if "occurred_at" in roles:
        score += 5
    if "amount" in roles:
        score += 5
    if "description" in roles:
        score += 3
    if "category_name" in roles:
        score += 2
    score += min(5, max(0, (len(table) - 1) // 5))
    return score, roles


# Las regex anchored (^/$) detectan formatos limpios. Las "first match"
# se usan para extraer la PRIMERA fecha de una celda con texto adicional
# (caso típico: "30/01/2026 Fecha valor 28/01/2026" en extractos donde
# la columna combina fecha operación y fecha valor).
_DATE_DDMM = re.compile(r"^(\d{1,2})/(\d{1,2})$")
_DATE_DDMMYY = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{2})$")
_DATE_DDMMYYYY = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")
_DATE_ISO = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")
_DATE_FIRST_DDMMYYYY = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")
_DATE_FIRST_DDMMYY = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2})\b")
_DATE_FIRST_DDMM = re.compile(r"\b(\d{1,2})/(\d{1,2})\b")
_DATE_FIRST_ISO = re.compile(r"(\d{4})-(\d{2})-(\d{2})")

_PERIOD_HINT = re.compile(
    r"(?:periodo|per[ií]odo|fecha emisi[oó]n|extracto)\D*?(\d{1,2})/(\d{1,2})/(\d{2,4})",
    re.IGNORECASE,
)


def _infer_year(full_text: str) -> int:
    """Busca en el texto del PDF un periodo o fecha de emisión.

    Si no encuentra ninguno, devuelve el año actual.
    """
    match = _PERIOD_HINT.search(full_text)
    if match:
        year_part = match.group(3)
        if len(year_part) == 2:
            return 2000 + int(year_part)
        return int(year_part)
    return date.today().year


def _normalize_date(value: str, default_year: int) -> str:
    """Normaliza una fecha del PDF a `DD/MM/YYYY` para que el service
    la parsee con sus formatos existentes. No lanza si no puede.

    Tolera celdas con texto extra: si la celda contiene varias fechas
    (típico en extractos que mezclan "Fecha operación + Fecha valor"
    en la misma columna, e.g. `"30/01/2026 Fecha valor 28/01/2026"`),
    se queda con la PRIMERA fecha que encuentra. La intención es la
    fecha operación: en columnas combinadas suele aparecer primera.
    """
    cleaned = (value or "").strip()
    if not cleaned:
        return ""
    # Match exacto preferido (celdas limpias).
    if _DATE_ISO.match(cleaned) or _DATE_DDMMYYYY.match(cleaned):
        return cleaned
    m = _DATE_DDMMYY.match(cleaned)
    if m:
        return f"{m.group(1)}/{m.group(2)}/20{m.group(3)}"
    m = _DATE_DDMM.match(cleaned)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{default_year}"

    # Fallback: extraer la primera fecha que aparezca dentro del texto
    # (orden de preferencia: ISO > DD/MM/YYYY > DD/MM/YY > DD/MM).
    m = _DATE_FIRST_ISO.search(cleaned)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = _DATE_FIRST_DDMMYYYY.search(cleaned)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    m = _DATE_FIRST_DDMMYY.search(cleaned)
    if m:
        return f"{m.group(1)}/{m.group(2)}/20{m.group(3)}"
    m = _DATE_FIRST_DDMM.search(cleaned)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{default_year}"
    return cleaned  # devolvemos como vino; el service decidirá si la rechaza


def parse_pdf_smart(payload: bytes) -> list[dict[str, str]]:
    """Heurística sobre `extract_tables`: detecta la tabla de
    transacciones entre todas las del PDF y devuelve filas con keys
    fijas (`amount`, `occurred_at`, `description`, `category_name`).

    El caller debe usar `VISION_FORCED_MAPPING` o equivalente cuando
    use el resultado de esta función — el mapping del usuario se
    ignora porque las claves ya son las correctas.

    Lanza `SmartParseAmbiguous` si ninguna tabla supera el score
    mínimo. Lanza `NoTablesInPdfError` si pdfplumber no encuentra
    ninguna tabla (igual que `parse_pdf`).
    """
    try:
        pdf = pdfplumber.open(io.BytesIO(payload))
    except Exception as e:
        # `str(e)` viene vacío para varias excepciones de PDFs cifrados
        # / corruptos. Caemos al tipo para que el usuario tenga al
        # menos una pista accionable en el log.
        detail = str(e) or f"{type(e).__name__} sin mensaje (probablemente PDF cifrado o corrupto)"
        raise ParseError(f"PDF inválido: {detail}") from e

    candidates: list[list[list[str | None]]] = []
    full_text_parts: list[str] = []
    with pdf:
        if len(pdf.pages) > _MAX_PDF_PAGES:
            raise ParseError(f"PDF demasiado largo (> {_MAX_PDF_PAGES} páginas).")
        for page in pdf.pages:
            try:
                text = page.extract_text() or ""
                if text:
                    full_text_parts.append(text)
            except Exception:
                pass
            for table in page.extract_tables():
                if table:
                    candidates.append(table)

    if not candidates:
        raise NoTablesInPdfError("No se detectaron tablas en el PDF (¿escaneado?)")

    # Agrupar tablas con cabeceras idénticas (continuación entre páginas).
    # Una vez identificada la "mejor" cabecera, todas las tablas posteriores
    # con esa misma cabecera se concatenan.
    scored = [
        (score, roles, table) for table in candidates for score, roles in [_score_table(table)]
    ]
    scored.sort(key=lambda x: x[0], reverse=True)

    best_score, best_roles, best_table = scored[0]
    if (
        best_score < _SMART_MIN_SCORE
        or "occurred_at" not in best_roles
        or "amount" not in best_roles
    ):
        raise SmartParseAmbiguousError(
            f"Heurística no encontró tabla de transacciones (mejor score={best_score})"
        )

    best_header_norm = [_norm(_pdf_clean(c)) for c in best_table[0]]
    grouped: list[list[list[str | None]]] = []
    for table in candidates:
        if not table:
            continue
        header_norm = [_norm(_pdf_clean(c)) for c in table[0]]
        if header_norm == best_header_norm:
            grouped.append(table)

    if not grouped:
        grouped = [best_table]

    inferred_year = _infer_year(" ".join(full_text_parts))

    rows: list[dict[str, str]] = []
    for table in grouped:
        for raw_row in table[1:]:  # saltar cabecera
            cleaned = [_pdf_clean(c) for c in raw_row]
            if all(not c for c in cleaned):
                continue
            # saltar repetidas de cabecera (algunos PDFs repiten la cabecera
            # al inicio de cada página)
            if [_norm(c) for c in cleaned[: len(best_header_norm)]] == best_header_norm:
                continue

            def get(role: str, cleaned: list[str] = cleaned) -> str:
                idx = best_roles.get(role)
                if idx is None or idx >= len(cleaned):
                    return ""
                return cleaned[idx]

            occurred_at = _normalize_date(get("occurred_at"), inferred_year)
            amount = get("amount")
            description = get("description")
            category_name = get("category_name")

            # Si Detalle está vacío usamos Concepto. Si Concepto y Detalle
            # son iguales (algunos extractos los duplican), nos quedamos
            # con uno solo.
            if not description:
                description = category_name
            elif description == category_name:
                category_name = ""

            rows.append(
                {
                    "amount": amount,
                    "occurred_at": occurred_at,
                    "description": description,
                    "category_name": category_name,
                    # PHASE-39 — saldo del extracto (vacío si no hay columna).
                    "statement_balance": get("statement_balance"),
                }
            )

    if not rows:
        raise SmartParseAmbiguousError("Tabla de transacciones detectada pero sin filas de datos")

    return rows
